#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 25 22:23:53 2019

@author: Thibaut Goldsborough, tg76@st-andrews.ac.uk
"""
#import tkinter as tk
#from tkinter import * 
from tkinter import filedialog

"""
tkinter is an inbuilt Python module used to create graphical user interfaces and widgets such as buttons, checkboxes, labels and menu bars.  tkinter is inbuilt in Python and does not need to be installed manually - hence why the importation code above is commented out. The workflow for using tkinter generally follows this structure:

Import tkinter modules --> create the main window for the app --> Add widgets --> enter the Main event loop.

The Main event loop is given by windowname.mainloop(). This command is given when the application is ready to run and it tells the code to keep displaying.
"""

import os, sys

"""
The module os is used to interact with the operating system of the device used, e.g. Windows, Mac, or Linux. The module sys is a set of functions which provide information about the interaction between Python and the operating system, such as what version of Python is running, and sys.exit(), which exits Python.
"""

basepath = filedialog.askdirectory(initialdir=os.path.dirname(__file__), title="Select folder to analyse")

"""
tkinter.filedialog is a module used to create file or directory selection windows.  Here, it is used to create a file selection window which asks the user to select the file containing the stacks of images that they wish to analyse using LECSegmenter.

askdirectory prompts the user to select a directory, i.e. the folder containing the stacks of images.

initialdir is the directory that the dialog starts in.  Here, os.path.dirname(__file__) returns a string value which is the directory name of the specified path, i.e. of the file path to the file that the user selected. Using __file__ as the argument of os.path.dirname() ensures that the initial directory, initialdir, is set to the directory of this Python script - i.e. the initial directory that is available for the user to select from is the folder where this script is saved, probably LECSegmenter#readme.

"""

print(basepath)

#Download the following packages: 
#%matplotlib auto

import numpy as np

"""
NumPy (Numerical Python) is an open source Python library that contains a comprehensive set of mathematical functions.  It is particularly useful for linear algebra.  NumPy provides N-dimensional array with indices starting at zero.
"""

#opencv
import cv2 as cv

"""
Opencv is an open source library for machine learning and image processing, among other applications.  Operations in opencv are combined with operations in NumPy to increase functionality.
"""

import PIL.Image

"""
PIL, or the Python Imaging Library, stores Python's image editing functions.  The class Image is used to represent an image within PIL.  
"""

from tkinter import * 

"""
import * is risky here.  import * imports all functions from the module tkinter.  If a user accidentally defines a function with the same name as a function within tkinter, the new definition will override the existing function from tkinter.  Therefore, the user should import tkinter as tk or similar, and call functions from tkinter using tk.functionname() when one wants to use the tkinter function and not a manually defined function.
"""

from tkinter import ttk 

"""
The module ttk provides 18 widgets for graphical user interfaces: Button, Checkbutton, Entry, Frame, Label, LabelFrame, Menubutton, PanedWindow, Radiobutton, Scale, Scrollbar, Spinbox, Combobox, Notebook, Progressbar, Separator, Sizegrip, and Treeview.  Importing ttk after 'from tkinter import *' means that ttk widgets override tk widgets (the older version of the widget module within tkinter).  The ttk.Style class is then used to style the widgets.  
"""

#from numpy import asarray
#from skimage.io import imsave
# from keras.preprocessing.image import array_to_img
# from keras.preprocessing.image import save_img
#skimage.morphology has several modules that can alter morphology of images-skeletanize-binary to one-pic

from skimage.morphology import skeletonize

"""
skimage (Scikit-image) is Python package used for image preprocessing.The module morphology contains functions that process images in different ways. For example, skeletonize(image) returns the skeleton of a binary image, whereas watershed(image, markers) returns a matrix labelled using the watershed segmentation algorithm.
"""


from random import randrange

"""
random.randrange returns a randomly selected element from the specified range.
"""

from matplotlib import pyplot as plt

"""
Plotting library.
"""


from scipy.signal import argrelextrema

"""
Scipy.signal is used for signal processing. argrelextrema(data, comparator[, axis, ...]) calculates the relative extrema of data. 'data' is the array in which to find the relative extrema. The comparator is the function used to compare two data points.  It should take two arrays as its arguments.
"""

import pickle as pickle

"""

"""


import easygui as gui

"""
Graphical user interface library.
"""

from openpyxl import Workbook 

"""
Openpyxl is a library for working with spreadsheet files compatible with Microsoft Excel. It allows one to read the spreadsheet, cycle through rows, extract data, and write new data to the spreadsheet. Workbook is the Excel compatible file modified; this may contain multiple worksheets. One can use Openpyxl to modify a Workbook without opening third-party Microsoft Excel software.

'import Workbook' creates a new Workbook with at least one sheet.  One can then name the Workbook using name = Workbook()
"""

global outcome, CELL_DICTIONARY, MAX_NUMBER_OF_CELLS, tiff_images, dim1, dim2

"""
Specifying the above variables to be global means that they have a global scope, as opposed to being defined for use within a specific function.
"""

MAX_NUMBER_OF_CELLS=0

"""
The global variable MAX_NUMBER_OF_CELLS is now assigned to zero.
"""

answer=gui.buttonbox("Would you like to Start Over or Continue Previous Project?",choices=("Start Over","Previous Project"))

"""
This opens a box asking you if you want to start over or continue.  Start Over means that you will be asked to select a file of image stacks to analyse.  If one chooses to Continue Previous Project, the programme will retrieve previously segmented images. 
"""


if answer=="Start Over":
    answer=True
    saving_location = filedialog.asksaveasfilename(initialdir = os.path.dirname(__file__),title = "Select saving location",filetypes = (("all files","*.*"),))
    
"""
If one chooses to start over, the answer is assigned the value True.  Then, asksaveasfilename, a function from the module filedialog within the module tkinter, produces a window directing the user to choose a saving location for the segmented images.  The initialdir is set to the folder in which this code is located - usually LECSegmenter#readme, which is called by __file__.
"""   
    
if answer=="Previous Project":
    answer=False
    loading_location = filedialog.askopenfilename(initialdir = os.path.dirname(__file__),title = "Select loading location",filetypes = (("all files","*.*"),))
    saving_location = loading_location
    
    """
    If the answer is 'Previous Project', the answer is assigned the value False.  A window is created which asks the user to select a loading location, that is, to select the file containing the previously analysed images.  The saving location is set to the loading location, such that, when changes to the images are made, the newer versions are saved to the same folder that they were retrieved from prior to modification.
    """

    with open(loading_location, 'rb') as infile:
        outcome = pickle.load(infile)
        
    """
    'with open(loading_location, 'rb') as infile:' opens the selected saved work in the reading and writing mode ('rb') as an infile, which means that the contents of this saved file will be eventually written out to a second file, the outfile.
    Serialization saves the state of an object from any process, so that it can be deserialisaed later to continue the process. Pickle is a Python-specific deserialization which converts the object into a binary string. Note that problems can occur if the file being unpickled was pickled using a different version of Python.
    Pickle.load unpickles the file infile.  Therefore, 'outcome' is the unpickled version of the file opened from the loading location.  In other words, the previous project is opened from the loading location chosen by the user, unpickled, and named 'outcome'.
    """
        
def merges_red(img1,img2,amount):
    overlay = cv.cvtColor(img1, cv.COLOR_GRAY2BGR)
    b,g,r = cv.split(overlay)
    r = cv.add(r,amount, dst = r, mask =img2, dtype = cv.CV_8U)
    merged=cv.merge((b,g,r),img1)
    return merged

"""
Defining a new function, merges_red, with arguments img1, img2, and amount. The purpose of this function is to produce the original image (img1) overlaid with a red skeleton outlining the cell membranes (img2). cvtColor is used to convert an image between colour spaces. It is accessed from Opencv (here cv2, imported as cv). cvtColor has syntax 
cv.cvtColor[src, code[,dst[,dstCn]])
where src is the name of the image to be converted, code is the colour space conversion code, and dst, dstCn are optional parameters. The code cv.COLOR_GRAY2BGR converts a grayscale image to blue-green-red channels.

b,g,r = cv.split(overlay) splits the overlay BGR image into separate blue, green, and red channels, which are independently assigned to the variables b,g, and r.

The variable r is then updated by adding 'amount' to the pixel values. dst is an output array that has the same size and number of channels as the input.  

cv.merge merges the split overlay image with img 1 to produce a single composite image, which is returned as the output of the function.

"""

#the same is repeated for blue and green channels
#merges_blue overlaps the blue labelled centroid labels in image 1 with the background in image 2
def merges_blue(img1,img2,amount):
    b,g,r = cv.split(img1)
    b = cv.add(b,amount, dst = b, mask =img2, dtype = cv.CV_8U)
    merged=cv.merge((b,g,r),img1)
    return merged
#green is eraser, merges_green overlaps the eraser in green with the background in image 2. 
def merges_green(img1,img2,amount):
    b,g,r = cv.split(img1)
    g = cv.add(g,amount, dst = g, mask =img2, dtype = cv.CV_8U)
    merged=cv.merge((b,g,r),img1)
    return merged

"""
The same is repeated for the blue and green channels.
"""


def removesmallelements(img,minsize):
    img=img.astype(np.uint8)
    nb_components, output, stats, centroids = cv.connectedComponentsWithStats(img, connectivity=8)
    sizes = stats[1:, -1]; nb_components = nb_components - 1
    img2 = np.zeros((output.shape))
    #for every component in the image, you keep it only if it's above minimum size of membrane
    for i in range(0, nb_components):
        if sizes[i] >=minsize:
            img2[output == i + 1] = 255
    return(img2)

"""
A function to remove small objects from the final image.  

img = img.astype(np.uint8) creates a copy of the image (that is, a copy of the image array) cast to the data type uint8, which stores 8-bit numbers (numbers 0-255).  If a value of an entry x is greater than 255, the returned value will be x mod 255.

cv.connectedComponentsWithStats(img, connectivity=8)
This function analyses an 8-bit single-channel (grayscale) image.  In this case, this is the input img, which was set to 8-bit using the previous function img.astype(np.uint8). The function searches for 8-way connectivity between pixels; that is, pixels are said to be neighbours if they touch at an edge or a corner (compare to 4-way connectivity, in which pixels are only classified as neighbours if they share an edge).
nb_components is the number of cells.  output labels each cell with a different number.  stats gives the surface area, perimeter, etc. of each cell.  Centroids gives the co-ordinates of the cell centroids.  The x- and y-coordinates of a cell centroid are extracted using centroids(label,0) and centroids(label,1) respectively.

sizes = stats[1:, -1]; nb_components = nb_components - 1
Defines the sizes of the membranes and reduces the number of cells by one.

img2 = np.zeros((output.shape))
Sets the pixels of img2 to zero.

The for loop ensures that only membranes above the specified minimum size are retained in the image.
"""


def nothing(x):
    pass

"""
This function is a placeholder.  When it is called, it is read by the interpreter and results in no operation.
"""


def watershed(img):
    global b1,stats
    edges= np.pad(np.ones((dim1-2,dim2-2)), pad_width=1, mode='constant', constant_values=0) 
    
    img=img*edges
    img=img.astype(np.uint8)
    
    """
    edges= np.pad(np.ones..) 
    
    np.pad returns a padded array of rank equal to an array with shape increased according to pad_width.  The array to be padded is np.ones((dim1-2,dim2-2)), which is an array of ones with dimensions two pixels less in each axis relative to img. The pad width is set to one pixel and the constant value of the pad is set to zero.  The output is an array the same size as img, with edge pixels of value 0 and inner pixels of value 1.
    
    This array is multiplied by the image array, such that all edge pixels in the image now have value 0 and all internal pixels are unchanged.
    
    The image is then cast to 8-bit type for consistency.
    """
 
    nb_components, output, stats, centroids = cv.connectedComponentsWithStats(cv.bitwise_not(img),\
    connectivity=4)
    
    """
    Extracts the number of cells, cell indices, statistics and centroid coordinates of the complement of img - that is, the image array that is analysed is the inversion of img, where pixels of value 1 in img now have value 0 and vice versa. 4-way connectivity.
    """

    b1=np.zeros((dim1,dim2)).astype(np.uint8)
    g1=np.zeros((dim1,dim2)).astype(np.uint8)
    r1=np.zeros((dim1,dim2)).astype(np.uint8)
    
    """
    Three zero arrays of the same dimensions as img, with 8-bit type, are created.
    """

    sizes = stats[1:, -1]; nb_components = nb_components - 1
    for i in range(0, nb_components):
        if sizes[i] <=10 :
            output[output==i+1]=0
            
    """
    The for loop removes components of total size less than or equal to 10 pixels from the watershed image by setting the output to zero if the size is <= 10 pixels.
    """

    output=cv.dilate(output.astype(np.uint8),None,iterations=1)
    
    """
    Dilation increases the area of the white regions of the image output.astype(np.uint8). This makes the cell boundaries thicker.  Increasing the number of iterations would increase the thickness of the white regions. Note that Opencv interprets white regions as the object, so a black boundary on a white background would become thinner under dilation.
    """

    for i in range(0, nb_components):
        if sizes[i] <=10000 : #MAX CELL SIZE
#[0] is for blue, [1] is for red and [2] is for green defined under colour[] later.
            b1[output == i + 1]=colors[i][0]
            g1[output == i + 1]=colors[i][1]
            r1[output == i + 1]=colors[i][2]
    image=cv.merge((b1,g1,r1))
    return (image,output)

"""
A random colour is assigned to each cell that has a size less than 10000 pixels.  This ensures that the area around the edges of the group of segmented cells, which is the background for our purposes but would be interpreted as a connected object, is not assigned a colour and remains black.

The zero arrays for blue, green and red channels (b1, g1 and r1) are modified so that pixels where the output array has the same value (i.e. pixels within the boundaries of a single cell) are assigned unique hues of blue, green and red using colors[i][0] (blue hue), colors[i][1] (green hue) and colors[i][2].  

The three colour channels are then merged to produce the watershed image, which is returned together with the cell outlines.
"""

def simple_watershed(img):
    global output,Centroid_list,nb_components
    Centroid_list=[]
    
    edges= np.pad(np.ones((dim1-2,dim2-2)), pad_width=1, mode='constant', constant_values=0)   
    img=img*edges
    img=img.astype(np.uint8)
    
    nb_components, output, stats, centroids = cv.connectedComponentsWithStats(cv.bitwise_not(img),\
    connectivity=4)
    
    sizes = stats[1:, -1]; nb_components = nb_components -1
 
    for i in range(0, nb_components):
        if sizes[i] <=10 :
            output[output==i+1]=0
           
    for cell_index in range(0,nb_components+1):
        if cell_index>1:
            if np.where(output==cell_index)[0].size>0:
                Centroid_list.append(((np.mean(np.where(output==cell_index)[0]),np.mean(np.where(output==cell_index)[1])),cell_index))
    #return the label on cell i.e. output and the centroid list with cx,cy coordinates
    return (output,Centroid_list)

"""
simple_watershed creates a list containing the (x,y) coordinates of each cell centroid and the index of the cell.  The function then returns the cell labels (output) and the list of centroid coordinates and cell indices.
"""


def get_centroids(frame):
    Centroid_list=[]
    nb_components=list(np.unique(frame))
    for cell_index in nb_components:
        if cell_index>1:
            if np.where(frame==cell_index)[0].size>0:
                Centroid_list.append(((np.mean(np.where(frame==cell_index)[0]),np.mean(np.where(frame==cell_index)[1])),cell_index))
    return (Centroid_list)


"""
frame is a mask in which cell pixels are labelled with the corresponding cell label.  nb_components=list(np.unique(frame)) creates a list of all the unique labels in frame, i.e. the indices of the cells.  
The for loop then considers cases where the cell index is greater than one and where the mean centroid x-coordinate (np.where(frame==cell_index)[0]) is greater than zero (probably for debugging purposes). Provided these two conditions are met, the cell index and the (x,y) co-ordinates of its centroid are appended to Centroid_list.
The output is a list of cell indices and corresponding centroid coordinates.
"""

def get_stats(input_data):
    global cnt,rect, CELL_DICTIONARY 
    background=np.zeros((dim1,dim2))
    CELL_DICTIONARY={}
    nb_components=list(np.unique(input_data[0][1]))

    """
    Define a function to obtain information about cell size, shape etc. that will be stored in CELL_DICTIONARY. First, we create a zero array (background) with the same dimensions as the image. 
    """

    for cell_index in range(MAX_NUMBER_OF_CELLS+1):
        if cell_index>1:
            CELL_DICTIONARY[cell_index]=[]

    for frame_index in range(len(input_data)): 
        frame=input_data[frame_index][1].copy()
        nb_components=list(np.unique(frame))
        
        for cell_index in nb_components:
            if cell_index>1:

           #take every cell within the frame with a cell index
                if np.where(frame==cell_index)[0].size>0:
                     #turn those cells to white in a black background; 255 represents white
                    background[frame==cell_index]=255
                    #find contours, maintain size of 8bit; 1 is the hierarchy/info on img topology;CHAIN_APPROX_NONEstores absolutely all the contour points
                    contours,hierarchy = cv.findContours(background.astype(np.uint8), 1,method= cv.CHAIN_APPROX_NONE)
                    #possibly inly drawing contours for the mebrane denoted 0
                    cnt = contours[0]
                    #cy,cx=coordinates of the centroid.
                    cy,cx=int(np.mean(np.where(frame==cell_index)[0])),int(np.mean(np.where(frame==cell_index)[1]))
                    #computes surface area within contour
                    area = cv.contourArea(cnt)
                    #computes closed contour perimeter/curve length
                    perimeter = cv.arcLength(cnt,True)
                    #finds the minimum-area bounding rotated rectangles for specified point sets. 
                    rect = cv.minAreaRect(cnt)
                    #finds the vertex of two diagonal points on the rectangle drawn around the cell
                    #rect0 and rect1 are the coordinates of the vertex
                    width=min(rect[1][1],rect[1][0]) #??
                    length=max(rect[1][1],rect[1][0])#???
                    #90 was added to get the angle to the horizontal rather than vertical
                    angle=rect[2]+90
                    if rect[1][1] > rect[1][0]:
                        #to prevent a flipped angle?
                        angle=rect[2]    
                    #if len(cnt)>5:
                      #  (x,y),(MA,ma),angle = cv.fitEllipse(cnt)
                      #draw an approximate rectangle around contours and get vertical and horizontal dimensions of rect which is more or less the dim of cell
                    vert_height = cv.boundingRect(cnt)[3]
                    #listing number of elements in the list with what exactly? what is [1]=cx? is it the x coordinate if centroid?
                    hoz_width=len(list(np.where(np.where(frame==cell_index)[1]==cx)[0]))
                    #gathering info on centroid coordinates, centroid coordinates, surface area, perimeter, etc for each cell and feeding it into the dictionary
                    CELL_DICTIONARY[cell_index].append(((cx,cy),int(area),int(perimeter),int(width),int(length),int(angle),vert_height,hoz_width))                              
                    #this resets the image to assign 0 to the whole image and runs the code again from background[frame==cell_index]=255 to set cell to white
                    background=np.zeros((dim1,dim2))
                    #if a new cell appears/an old cell disappears in a movie the set NA if information is not available on area, centroids etc 
                    # this only apples to video and not images since all cells appear at the same time in the same image
        for cell_index in range(MAX_NUMBER_OF_CELLS+1):
            if cell_index>1:
                if len(CELL_DICTIONARY[cell_index])<=frame_index:
                    
                    CELL_DICTIONARY[cell_index].append((("NA","NA"),"NA","NA","NA","NA","NA","NA","NA"))
           #?        
    for cell_index in range(MAX_NUMBER_OF_CELLS+1):
        if cell_index>1:
            if all(elem ==(('NA', 'NA'), 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA') for elem in CELL_DICTIONARY[cell_index])==True:
                del(CELL_DICTIONARY[cell_index])
                
        
    return (CELL_DICTIONARY)



#suppose we have two images img 1/pre-image at time t and img 2/post-image at time t+1
def follow_cells_and_watershed(prev_img,img):
    #this function tries to go through every frame in the video and tries to match a cell in the post image with the pre-image
    global pairs_of_cells,mask,MAX_NUMBER_OF_CELLS
    # we getting centroid coordinates from previous image/img2 and trying to match it to the centroid in the pre-image/image1
    centroids_prev=get_centroids(prev_img)
    post_output,centroids_post=simple_watershed(img)
    pairs_of_cells=[]
    for centroid1num in centroids_prev:
        dist_list=[]
        #go through every centroid in first image
        centroid1=centroid1num[0]
        for centroid2num in centroids_post:
            #do through every centroid in the second image;#0 represents both the x,y coordinates (check get_stats: centroid list)
            centroid2=centroid2num[0]
            #calculate the distance between the coordinates of centroids in img 1 and centroids in img 2;
            #the 0 here represents the first element of x,y which is x coordinate and 1 represents second element which is y coordinate
            #(centroid1num[1],centroid2num[1])) denotes the centroids itseld
            dist_list.append((np.sqrt((centroid1[0]-centroid2[0])**2+((centroid1[1]-centroid2[1])**2)),(centroid1num[1],centroid2num[1])))
            #if the distance between the centroids in the pre image and the post image is <50 i.e. the closest distance then assign it as the same cell i.e. tracked cell from img 1 to img
          #because dist list is of the form [distance, x coordinate, y coordinate] 0 below represents the first element in the list 
        if min(dist_list, key = lambda t: t[0])[0]<50:
            #i.e distance and if dist <50 track as same cell i.e pair the cell in img 1 with img2 as the same cell but at a different location
            pairs_of_cells.append((min(dist_list, key = lambda t: t[0])[1],min(dist_list, key = lambda t: t[0])[0]))
    mask=np.zeros((dim1,dim2))
    #suppose cell 2 in img 1 is same as cell 6 in img 2, post_cell_nums=6;pre_cell_nums=2
    
    post_cell_nums = [lis[0][1] for lis in pairs_of_cells] ##??
    pre_cell_nums= [lis[0][0] for lis in pairs_of_cells]
    #MAX_NUMBER_OF_CELLS in pre- and post image
    MAX_NUMBER_OF_CELLS=max(MAX_NUMBER_OF_CELLS,max(post_cell_nums),max(pre_cell_nums))
    #unique numbers for every single cell in the output
    for cell_index in np.unique(post_output):
        if cell_index>1:
            #if there is a new cell in the post img and is not found in pre-img then add 1 to the max number of cells since it means there is a new cell
            if (cell_index in post_cell_nums)==False:
                MAX_NUMBER_OF_CELLS+=1
                pairs_of_cells.append(((MAX_NUMBER_OF_CELLS,cell_index),1000))
     #If the program detects two cells in img 2 which both have centroids close to the cell in img 1
    cell_num=len(pairs_of_cells)
    for pair_index in range(cell_num):
        pair=pairs_of_cells[pair_index]
        for pair2_index in range(cell_num):
            pair2=pairs_of_cells[pair2_index]
            if pair2[0][1]==pair[0][1]:
                if pair2[0][0]!=pair[0][0]:
        #then identify the cell from the two that has the smallest distance 
                    pairs_of_cells[pair_index]=pairs_of_cells[pair2_index]=min((pair,pair2), key = lambda t: t[1])
        #mask[post_output==pair[0][1]]=pair[0][0]
    
    for pair in pairs_of_cells:
        #post_output is the frame with all cell numbers
        #ensure the paired cells in img 1 and 2 get the same cell number
        mask[post_output==pair[0][1]]=pair[0][0]
            
    mask=cv.dilate( mask.astype(np.uint8),None,iterations=1)
    #assign zeros to the membrane
    b2=np.zeros((dim1,dim2)).astype(np.uint8)
    g2=np.zeros((dim1,dim2)).astype(np.uint8)
    r2=np.zeros((dim1,dim2)).astype(np.uint8)
    #assign a unique color to each cell
    for i in (np.unique( mask).astype(np.uint8)):
        if i>1:
            b2[ mask == i ]=colors[i-1][0]
            g2[ mask == i ]=colors[i-1][1]
            r2[ mask == i  ]=colors[i-1][2]
    image=cv.merge((b2,g2,r2))
    #imgge is where every pixel has 3 values (RGB) and mask is an image where each pixel has 1 value
    return image, mask




def GUI(event,x,y,flags,param):
    
    global Skeletonized_Image,Cursor
    global drawing, X,Y ,been
    global saved_list,dim, iter_photo
    
    Cursor=np.zeros((dim1,dim2)).astype(np.uint8)
    Skeletonized_Image=saved_list[len(saved_list)-1].copy()

"""
Function to create the graphical user interface.  The cursor is first created as an 8-bit zero array with the same dimensions as the image. 
"""

#drawing mode
    if drawing==True:
        if mode==True:
            if event == cv.EVENT_LBUTTONDOWN:
                #cv.line helps draw lines in drawing mode
               debug = cv.line(Skeletonized_Image,(x,y),(X,Y),(255),1).copy()
    #each time you modify image you want to save the changes and update the cell numbers each time
               saved_list.append(debug)
               update_numbers(saved_list[-1],iter_photo)
               drawing=False 
               been=False
   #if you click once don't draw a line but if you click the seond time draw a line betweem the forst click and the second            
    if been ==True:
        if drawing==False:
            if mode==True:
                if event == cv.EVENT_LBUTTONDOWN:
                    drawing=True
                    X,Y=x,y               
    if drawing==False:
        been=True
#this is the cursor mode which helps erase; cv.circle helps draw a circle for eraser
    if mode==False:
        cv.circle(Cursor,(x,y),dim, (1), 0)
        cv.circle(Skeletonized_Image,(x,y),dim, (0), -1)
        if event== cv.EVENT_LBUTTONDOWN:            
            saved_list.append(Skeletonized_Image.copy())
            update_numbers(saved_list[-1],iter_photo)
            
            
            

def process_image(img,a,b,c,d):
    if c <1:
        c=1
    #Note: 
 #cv.GaussianBlur blurs the image and gives the first image in GUI sequence.  
    BLURED= cv.GaussianBlur(img,(5,5),0)
#adaptiveThreshold gives the whole image a uniform luminosity e.g if some parts of the image are too bright it will lower the luminosity and conserse is true
#Blurred is the input img, 255 is max value of pixel for which condition is satisfied;. ADAPTIVE_THRESH_MEAN_C transforms a grayscale image to a binary image and c represents the Adaptive threshold toggle on the GUI image 1.
#The output of this is the second image in the GUI sequence of images being displayed. 
    GAUSSTHRESH=cv.adaptiveThreshold(BLURED,255,cv.ADAPTIVE_THRESH_MEAN_C,cv.THRESH_BINARY,2*c+1,0)  
  #This function removes small blobs and gives the 3rd image in the sequence of images in GUI. 
    rem=removesmallelements(GAUSSTHRESH,1000)
    img3 = cv.GaussianBlur(rem,(5,5),0)
    #Another threhold where img3 =source img, 255 is max non-zero value to pixel for which condition is satisfied
    #cv.THRESH_BINARY= type of threhold that transforms a grayscale image to a binary image. a represents the second threshold. 
    #The result of this is the 4th image in the sequence. 
    ret,img4 = cv.threshold(img3,a,255,cv.THRESH_BINARY)
    #again remove any small blobs 
    rem=removesmallelements(img4,1000)
    #Do another round of blurring
    img5 = cv.GaussianBlur(rem,(5,5),0)
    #cv.threshold/the third threshold. If above a certain threhold lower it and if below a certain threhold increase it.
    #Note: d is not defined. 
    ret,img6 = cv.threshold(img5,b,255,cv.THRESH_BINARY)
    Skeletonized_Image = (skeletonize(img6//255) * 255).astype(np.uint8)
    Watershed=watershed(Skeletonized_Image)[0]  
    img6=cv.cvtColor(img6.astype(np.uint8), cv.COLOR_GRAY2BGR)
    #this gives a red skeleton to the membrane which is seen in image 5 in GUI
    Skeletonized_Image_BGR=merges_red(img//2,Skeletonized_Image,255)
 #Next, the skeletonkized image is converted to a watershed and this is image 6 in the sequence of images. 
    return (img,GAUSSTHRESH,img4.astype(np.uint8),img6,Skeletonized_Image_BGR,Watershed,Skeletonized_Image)





#There are two modes but slow was not used probably because it was too slow.
#update numbers is used to update the cell numbers each time you draw of the image in drawing mode
def update_numbers(membrane_outlines,frame_num,speed="Fast"):
    global Numbers, tiff_images
    Image=tiff_images[frame_num]
    Base_Image=2*(Image[0]+Image[1])
    
    Numbers=np.zeros((dim1,dim2)).astype(np.uint8)
    if speed=="Slow":  
        mask=outcome[frame_num][1].copy()        
        centroids=get_centroids(mask)
        
    if speed=="Fast":
        mask=watershed(membrane_outlines)[1]
        #this retrives the centroids of every cell 
        centroids=get_centroids(mask)
               
    for centroid in centroids:
    #Add text i.e. add cell numbers as blue labels at the coordinates of the centroid
        cv.putText(Numbers,str(centroid[1]),org=(int(centroid[0][1]-5),int(centroid[0][0]+5)),fontFace=cv.FONT_HERSHEY_SCRIPT_SIMPLEX,fontScale=0.3,thickness=0,color=(1))
  #the red outline is assigned to the membrane
    pre_left_panel=merges_red(Base_Image//2,membrane_outlines,255)
    #the blue is assigned to the labels of cell numbers in the middle of the cell
    left_panel=merges_blue(pre_left_panel,Numbers,255)

    return left_panel




def save_all_work(boolean):
    print('Start saving')
    global outcome
    for frame_num in range(len(outcome)):
#For every frame in a video, outcome has 3 diff values where channel 3 is the color image with 3 different channels.
#Channel 1 is the one channel image with red memrbanes; Numbers_mask is the image with text numbers in blue for each cell.  
        outline=outcome[frame_num][2].copy()    
        if len(outcome[frame_num][2])==0:
              break
#This is applicable to an image with 1 frame. The changes are saved into outcome. 
        if frame_num==0:
            Channel3_Mask,Channel1_Mask=watershed(outline)
            outcome[frame_num][1]=Channel1_Mask.copy()
            outcome[frame_num][0]=Channel3_Mask.copy()
            Numbers_Mask=update_numbers(outline,frame_num,"Slow")
            outcome[frame_num][3]=Numbers_Mask.copy()              

#This is applicable to a video with several frames. The changes are saved into outcome. 

        if frame_num>0:
            prev_frame=outcome[frame_num-1][1].copy()
            Channel3_Mask,Channel1_Mask=follow_cells_and_watershed(prev_frame,outline)
            outcome[frame_num][1]=Channel1_Mask.copy()
            outcome[frame_num][0]=Channel3_Mask.copy()
            Numbers_Mask=update_numbers(outline,frame_num,"Slow")
            outcome[frame_num][3]=Numbers_Mask.copy()
            
        for image_index in range(len(outcome[frame_num])):
            image_to_save = PIL.Image.fromarray(outcome[frame_num][image_index])
            image_to_save.save(saving_location + '_im_' + str(image_index) + '_' + str(frame_num) + '.tiff')
        # save the segmentation frame here:
#               saving_location ='saved_work'

    if boolean==True:
        with open(saving_location, 'wb') as outfile:
            pickle.dump(outcome, outfile, pickle.HIGHEST_PROTOCOL)
    
    if frame_num > 1:
        CELL_DICTIONARY=save_excel(outcome,saving_location + '_cell_info.xls')
                  #outcome is a tuple so tuple to np array
#                   outcome2= np.asarray(outcome).astype(np.unit8)
                  #array to img 
#                   outcome3 = array_to_img(outcome3)
#                   save_img('test.tif', outcome3)     
    print('Finished saving stuff.')
        
         
            
            
            
   #This allows you to visualize the changes you have made but editing is not possible here.
#Note: this definition is linked to another file called visalize.
def display(outcome):
    cv.namedWindow('Press E to Exit')  
    frame_num=0  
    while(1):
        k = cv.waitKey(1) & 0xFF
        display_right=outcome[frame_num][0]
        display_left=outcome[frame_num][3]
    #hstack is used to display two images side by side
        display= np.hstack((display_left,display_right))
        cv.imshow('Press E to Exit',display)
        cv.moveWindow('Press E to Exit',150,10)
        #go to next frame
        if k ==ord('p'):
            if frame_num<(len(outcome)-1):
                frame_num+=1
        #go to previous frame
        if k==ord('o'):
            if frame_num>0:
                frame_num-=1
        #exit
        if k==ord('e'):
            break              
        
    cv.destroyAllWindows()
    
    
    
#The following def saves all work to an excel file known as Cell info (mentioned later in the code).
def save_excel(outcome,Save_As):
    wb = Workbook()
#we add a sheet in excel and only work with 1 sheet i..e there are not additional sheets
    sheet1 = wb.create_sheet('Sheet 1')
#we retrive the statistics including surface area, centroid coordinates etc from the cell_dictionary
    CELL_DICTIONARY=get_stats(outcome)
#The number of frames will define the number of tiff images
    number_of_frames=len(tiff_images)
#Add a title called cell number in the first row first column
    sheet1.cell(1,1,'Cell Number')
#Add a title called Parameter in the first row second column
    sheet1.cell(1,2,'Parameter')
    row=1
    for i in range(number_of_frames):
        #TODO consider break statement here
        sheet1.cell(1,3+i,'Frame'+str(i))

    row=-8
    cell_index2=0
    for cell_index in list(CELL_DICTIONARY.keys()):
        row+=8
    #Define the coordinates in teh excel sheet where you want the following labels.
    #The +8 means that the labels are added after every 8th row sunh that 
    #the first 8rows represent the parameters of the first cell the second set of 8 rows the next ect.
        sheet1.cell(2+cell_index2*8,2,"Centroid")
        sheet1.cell(3+cell_index2*8,2,"Surface Area")
        sheet1.cell(4+cell_index2*8,2,"Perimeter")
        sheet1.cell(5+cell_index2*8,2,"Width")
        sheet1.cell(6+cell_index2*8,2,"Length")
        sheet1.cell(7+cell_index2*8,2,"Angle")
        sheet1.cell(8+cell_index2*8,2,"Vertical Height")
        sheet1.cell(9+cell_index2*8,2,"Horizontal Width")
        cell_index2+=1
        for j in range(8):        
            sheet1.cell((j+1)+row+1,1,str(cell_index))
        for frame in range(number_of_frames):
#The following uses the information in the cell dictionary to fill out the stats information in the excel sheet
            sheet1.cell(1+row+1,3+frame,str((CELL_DICTIONARY[cell_index][frame][0]))) #2nd row, 2nd column has centroid info
            sheet1.cell(2+row+1,3+frame,str((CELL_DICTIONARY[cell_index][frame][1]))) #3rd row, 2nd column has surface area info
            sheet1.cell(3+row+1,3+frame,str((CELL_DICTIONARY[cell_index][frame][2]))) #4th row, 2nd column has perimeter info
            sheet1.cell(4+row+1,3+frame,str((CELL_DICTIONARY[cell_index][frame][3]))) #5th row, 2nd column has width
            sheet1.cell(5+row+1,3+frame,str((CELL_DICTIONARY[cell_index][frame][4]))) #6th row, 2nd column has length
            sheet1.cell(6+row+1,3+frame,str((CELL_DICTIONARY[cell_index][frame][5]))) #7th row, 2nd column has angle
            sheet1.cell(7+row+1,3+frame,str((CELL_DICTIONARY[cell_index][frame][6]))) #8th row, 2nd column has vertical height
            sheet1.cell(8+row+1,3+frame,str((CELL_DICTIONARY[cell_index][frame][7]))) #9th row, 2nd column has the horizontal width
            
#create a new file and name i.e. Cell_info.xls      
    wb.save(Save_As)
    
    return CELL_DICTIONARY
#not implemented in program



def plot_cell_movement(CELL_DICTIONARY):     
    X = np.array(())
    Y= np.array(())
    U = np.array(())
    V = np.array(())
    for cell_index in list(CELL_DICTIONARY.keys()):
        if type(CELL_DICTIONARY[cell_index][-1][0][1])!=str:
            if type(CELL_DICTIONARY[cell_index][0][0][1])!=str:        
                startx=CELL_DICTIONARY[cell_index][0][0][1]
                starty=CELL_DICTIONARY[cell_index][0][0][0]
                X=np.append(X,startx)
                Y=np.append(Y,starty)        
                endx=CELL_DICTIONARY[cell_index][-1][0][1]
                endy=CELL_DICTIONARY[cell_index][-1][0][0]
                vectX=(endx-startx)
                vectY=(endy-starty)    
                U=np.append(U,vectX)
                V=np.append(V,vectY)  
    fig, ax = plt.subplots()
    ax.quiver(X, Y,U,V,units='xy' ,scale=0.2,headwidth=2)   
    plt.grid()    
    ax.set_aspect('equal')  
    plt.xlim(0,dim1)
    plt.ylim(0,dim2)  
    plt.show()

    
    
    
    
    
    
    
    
    
    
    
    
    
    
#Definitions complete. 
#This is the start of the code.  
print("Reading files...")

#Basepath is the path to the stacks folder withj the images.
photos=[]
for entry in os.listdir(basepath): #Read all photos
    if os.path.isfile(os.path.join(basepath, entry)):
    #creating a list of every image in photos
        photos.append(entry)
#to sort the images numerically/alphabetically
photos.sort()

list_of_means=[]

#Estimate the mean brightness and only select the brightest stack and the stack before the brightest
for tiff_index in range(len(photos)): 
    if photos[tiff_index]!='.DS_Store':
        print('reading in the file')
        print(photos[tiff_index])
        tiff_photo=cv.imread(basepath+"/"+photos[tiff_index])
#mean intensity of the pixels/mean intensity of the image
        list_of_means.append(np.mean(tiff_photo))
#if you plot lis_of_means you find distinct peaks in intensity of each image studied 
dim1=np.shape(tiff_photo)[0] 
dim2=np.shape(tiff_photo)[1] 
    
array_of_means=np.array(list_of_means)   
#find the maximum and minimum intensity from the list_of_means
local_maxima=argrelextrema(array_of_means, np.greater)[0]
local_minima=argrelextrema(array_of_means, np.less)[0]

false_maximas=[]
#there are 2 maximas for each img (1 true and another false maxima) which can be visualised with plt.plot(list_of_means)
local_maxima_list=[]
for maxima in local_maxima:
    local_maxima_list.append(maxima)
    
#check if the local maxima is actually a local minima
for minima in local_minima:
    for maxima in local_maxima:
        if minima==maxima+1:
            false_maximas.append(maxima)
#remove all the false maximas          
for false in false_maximas:
    local_maxima_list.remove(false)

false_maximas=[]
for maxima_index in local_maxima_list:
    if list_of_means[maxima_index]<=np.mean(list_of_means):
        false_maximas.append(maxima_index)
#only give the index i.e. img numbers of true maxima in list local_maxima_list.remove
#Note of nearly 160 images, the maxima were found at nearly 16, 39, 60,90,110,130 and 150 where numbers mean the img numbers
for false in false_maximas:
    local_maxima_list.remove(false)
        
print("Done")


#run through the local maxima images, and take each img with local maxima/most well defined img and add the previous image to it
tiff_images=[]
for Image in local_maxima_list:
    print('and now rereading the image')
    print(photos[Image-1])
    tiff_images.append((cv.imread(basepath+"/"+photos[Image],cv.IMREAD_GRAYSCALE),(cv.imread(basepath+"/"+photos[Image-1],cv.IMREAD_GRAYSCALE))))   

Info_Sheet=cv.imread("./Info_Sheet.tiff",cv.IMREAD_GRAYSCALE)
Info_str="""

                                Info Sheet
        
 Commands:   
    "p" --> Go to next frame
    "o" --> Go back one frame
    "l" --> Switch between Manual segmentation and Automatic segmentation
    "s" --> Save
    
 User Interface:   
    "m" --> Switch between Drawing/Erasing
    "z" --> Undo
    
    Drawing:   
        Click at two points to draw a line segment
    
    Erasing:    
        "b" --> Increase diameter of eraser
        "n" --> Decrease diameter of eraser
        

    If you have any questions or comments please email tg76@st-andrews.ac.uk

"""
#define all colours
colors=[]
for i in range(10000):
    colors.append((randrange(255),(randrange(255)),(randrange(255))))

#if display 
iter_photo=0
Display_Mode=False
#restart a new project with new outcome i.e. black outcome

if answer==True:
    outcome=[]
    for i in range(len(tiff_images)):
        outcome.append([])
        for j in range(4):
            outcome[i].append([])


dim=1 ##??
saved_once2=False
#image refers the two best images i.e. ones with highest resolution/definition;
#image [0] is the one with highest local maxima and image[1] is the one before image [0]
while iter_photo < len(tiff_images):

    Cursor=np.zeros((dim1,dim2)).astype(np.uint8)
    Numbers=np.zeros((dim1,dim2)).astype(np.uint8)
    Image=tiff_images[iter_photo]
    Image_str='Image'+str(iter_photo+1) 
    #img=4*np.maximum(Image[0],Image[1])
    #was multipled by 2 to make the image brighter
    Base_Image=2*(Image[0]+Image[1])
    #start with a clean sheet
    cv.destroyAllWindows()
    Image_str="GUI "+Image_str
    lag_iter=iter_photo
    mode = True  
    drawing=False 
    been=False
    
    saved_list=[]
    saved_list.append(1)
    #answer=false means we want to continue the previous project and this was saved under a variable called outcome
    if answer==False:
        print('Attempting to obtain skeletonization from saved data')
        Skeletonized_Image=outcome[iter_photo][2]
        # if this frame was not processed in previous iterations, then we won't
        # actually have produced a skeletonized image yet, and have to generate it
        # from scratch, hence the next two lines.
        if len(Skeletonized_Image)==0:
            print('Skeletonization not available from saved data, so I will make a new one')
            Skeletonized_Image=frame=process_image(Base_Image,118,53,5,1)[6]
        print('Done. We have a skeletonized image for this frame now')
    #if answer=true then we choose to start a new project and it runs through the process_image defined above  
    if answer==True:
        Skeletonized_Image=frame=process_image(Base_Image,118,53,5,1)[6]
    #Skeletonized_Image is the image with membranes outlined 
    saved_list.append(Skeletonized_Image)
    #defining a popup window; Info Sheet for how the program works
    cv.namedWindow("Info Sheet",flags=cv.WINDOW_NORMAL)
    #location of window defined by 500,10
    cv.moveWindow("Info Sheet",500,10)
    #dimensions of the window defined by 600,500
    cv.resizeWindow("Info Sheet",600,500) 

    #This is the popup that allows you to manually edit the two images
    if Display_Mode==False:      
        cv.namedWindow(Image_str,flags=cv.WINDOW_NORMAL)
    #location of the window on the screen
        cv.moveWindow(Image_str,10,10)
    #dimensions of the window
        cv.resizeWindow(Image_str,1300,800)
    #To create switches labelled contours and labels for the trackbar
        switch = 'Contours: 0 : OFF \n1 : ON'
        switch2 = 'Labels: 0 : OFF \n1 : ON'
        cv.createTrackbar(switch, Image_str,1,1,nothing)
        blablatest = cv.getTrackbarPos(switch,Image_str)
        print("the trackbar position is now:")
        print(blablatest)
        cv.createTrackbar(switch2, Image_str,1,1,nothing)
        cv.setMouseCallback(Image_str,GUI)
  #This is the popup with 6 images that allows you to alter different threholds but does not allow drawing 
    if Display_Mode==True: 
        cv.namedWindow(Image_str,flags=cv.WINDOW_NORMAL)
    #location of the window on the screen defined by 10,10
        cv.moveWindow(Image_str,10,10)
     #dimensions of the window defined by 1300,800
        cv.resizeWindow(Image_str,1300,800)
    #label the trackbar for threholds
        cv.createTrackbar('Second Threshold',Image_str,118,200,nothing)
        cv.createTrackbar('Third Threshold',Image_str,53,200,nothing)
        cv.createTrackbar('First (Adaptive) Threshold',Image_str,5,20,nothing)
        cv.createTrackbar('On/Off',Image_str,1,1,nothing)
    #update the cell numbers each time you move to the next image/when the screen pops up for the first time initialize all numbers
    update_numbers(saved_list[-1],iter_photo)
    
#creating a loop until break is used
    while(1):
        if not Display_Mode:
            a1 = cv.getTrackbarPos(switch,Image_str) #this is contours on the trackbar
            a2 = cv.getTrackbarPos(switch2,Image_str) #this is labels on the trackbar
        k = cv.waitKey(1) & 0xFF
        #have you clicked on button m? 
        if k == ord('m'):
            mode = not mode
        #have you clicked on button p? if yes, go to the next photo. 
        if k == ord('p'):
            lag_iter=iter_photo
            iter_photo+=1           
            break 
        if k == ord('o'):
         #have you clicked on button 0? if yes, go to the previous photo. 
            if iter_photo>0:
                lag_iter=iter_photo
                iter_photo-=1
                break 
        if k == ord('b'):
        #if clicked on b, increase the diameter of the circle by 5 pixels 
            dim+=5
        if k == ord('n'):
        #if clicked on n, decrease the diameter of the circle by 5 pixels 
#need to check if dim is >5 to prevent negative diameter
            if dim>=5:
                dim-=5
            else:
                dim=1
        if k== ord('z'):
        #if clicked on z: then undo the last element of the list 
           if len(saved_list)>2:
               #delete the last element of the list
               del saved_list[-1]
               print(saved_list)
               #reinitialise the cell numbers on the new list i.e without the element that was deleted
               update_numbers(saved_list[-1],iter_photo)
    #switches to display mode 
        if k==ord('l'):
            Display_Mode=not Display_Mode
            break
    #s is to save all the work 
        if k==ord('s'):
            watershed_Image=watershed(saved_list[-1])[0]
        #every time you make a modification it will not be saved, Only the last element in saved list will be saved to saved_list[-1]
            outcome[iter_photo]=(list((saved_list[-1],watershed_Image,saved_list[-1],update_numbers(saved_list[-1],iter_photo))))
            
            save_all_work(True)
            
#This is the mannual mode with two images 
        if Display_Mode==False:
#take saved_list[-1] is watershed i.e. there is a color to every cell
            Watershed=watershed(saved_list[-1])
#The right panel will have the watershed with all 3 channels denoted by Watershed[0] 
            right_panel=Watershed[0]
#on the left panel we have the two images with the highest resolution which was initially multipled by 2 so here we divide by 2 again
#make all the membranes red where a1*255 is on the other end of the contours switch on trackbar
            pre_pre_left_panel=merges_red(Base_Image//2,saved_list[-1],a1*255)
#make all the membranes red where a2*255 is on the other end of the labels switch on trackbar
            pre_left_panel=merges_blue(pre_pre_left_panel,Numbers,a2*255)
#this allows the left panel to have a green cursor 
            left_panel=merges_green(pre_left_panel,Cursor,255)
#hstack horizontally stacks the two panels aside each other
            window= np.hstack((left_panel,right_panel))
            if saved_once2==True:
                saved_once2=False
                saved_list.append(frames[6])
                update_numbers(saved_list[-1],iter_photo)
#This is the mode where you can only change threholds but not mannual editing
        if Display_Mode==True: 
#These are the trackbars for the automatic popup
            a2 = cv.getTrackbarPos('Second Threshold',Image_str) 
            b2 = cv.getTrackbarPos('Third Threshold',Image_str) #53
            c2= cv.getTrackbarPos('First (Adaptive) Threshold',Image_str)
            d2 = cv.getTrackbarPos('On/Off',Image_str) #no particular fucntion
#calls the function process_image defined earlier to process the images in this popup         
            frames=process_image(Base_Image,a2,b2,c2,d2)
#this horizontally stacks 3 images one beside each other in the first row of images
            horizontal1 = np.hstack((frames[0],frames[1],frames[2]))
            bgrhorizontal1 = cv.cvtColor(horizontal1.astype(np.uint8), cv.COLOR_GRAY2BGR)
#this horizontally stacks 3 images one beside each other in the second row of images
            bgrhorizontal2 = np.hstack((frames[3],frames[4],frames[5]))
#this vertically stacks images in both rows one on top of each other
            window=np.vstack((bgrhorizontal1,bgrhorizontal2))
#reinitialise the saved work everytime you switch modes?       
            saved_once2=True
#show windowss
        cv.imshow(Image_str,window)
#show info sheet on how program works
        cv.imshow("Info Sheet",Info_Sheet)
#loop is done
    cv.destroyAllWindows()
#create the watershed image, and sace it in outcome
    watershed_Image=watershed(saved_list[-1])[0]
    outcome[lag_iter]=(list((saved_list[-1],watershed_Image,saved_list[-1],update_numbers(saved_list[-1],lag_iter))))
    print(outcome[lag_iter])
print("Processing....")
#popup for save
answer=gui.buttonbox("Would you like to save this?",choices=("Save","Don't Save"))
#this calls the definition save_all_work created earlier and if boolean is true then dump it into a pickle
if answer=="Save":
    answer=True
if answer=="Don't Save":
    answer=False

#sys.exit(0)
#save work if answer=true 
save_all_work(answer)

display(outcome)
#save to excel file called Cell_Info with all the info in cell dictionary
CELL_DICTIONARY=save_excel(outcome,'Cell_Info.xlsx')

print("Done")

#To obtain a matplotlib plot of the cell movements
#plot_cell_movement(CELL_DICTIONARY)








